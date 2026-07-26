# ============================================================
# 00. INSPECT THE OFFICIAL 2024-25 HES APC PUBLIC FILES
# ============================================================
# Downloads the final annual summary, ICB, provider and provider-level-analysis
# files. The output records workbook structure and distinct CSV dimensions so the
# production extractor can use published total rows rather than overlapping sums.

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import requests


FILES = {
    "summary": "https://files.digital.nhs.uk/EA/F24476/hosp-epis-stat-admi-summary-tabs-2024-25.xlsx",
    "icb": "https://files.digital.nhs.uk/B0/CC3BB5/hosp-epis-stat-admi-icb-resp-2024-25-tab.xlsx",
    "provider": "https://files.digital.nhs.uk/DF/EE7338/hosp-epis-stat-admi-hosp-prov-2024-25-tab.xlsx",
    "provider_level_analysis": "https://files.digital.nhs.uk/3C/145B13/hosp-epis-stat-admi-pla-2024-25.csv",
}
OUTPUT_PATH = Path("public-data/hes-apc-2024-25-inspection.json")


# ============================================================
# 01. NORMALISE CELL VALUES FOR JSON
# ============================================================
def json_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value,"isoformat"):
        return value.isoformat()
    if isinstance(value,(str,int,float,bool)):
        return value
    return str(value)


def nonempty_rows(worksheet,max_rows: int = 40) -> list[dict[str,Any]]:
    rows = []
    for row_number,row in enumerate(worksheet.iter_rows(values_only=True),start=1):
        values = [json_value(value) for value in row]
        if not any(value not in (None,"") for value in values):
            continue
        rows.append({"row":row_number,"values":values[:40]})
        if len(rows) >= max_rows:
            break
    return rows


# ============================================================
# 02. INSPECT EXCEL WORKBOOKS
# ============================================================
def inspect_workbook(name: str,url: str) -> dict[str,Any]:
    response = requests.get(url,timeout=300)
    response.raise_for_status()
    workbook = openpyxl.load_workbook(io.BytesIO(response.content),read_only=True,data_only=True)

    sheets = []
    for worksheet in workbook.worksheets:
        sheets.append({
            "name":worksheet.title,
            "max_row":worksheet.max_row,
            "max_column":worksheet.max_column,
            "preview":nonempty_rows(worksheet),
        })

    return {"name":name,"url":url,"bytes":len(response.content),"sheet_names":workbook.sheetnames,"sheets":sheets}


# ============================================================
# 03. INSPECT PROVIDER-LEVEL ANALYSIS CSV
# ============================================================
def inspect_csv(name: str,url: str) -> dict[str,Any]:
    response = requests.get(url,timeout=300)
    response.raise_for_status()
    frame = pd.read_csv(io.BytesIO(response.content),dtype=str,low_memory=False,encoding="utf-8-sig")

    distinct = {}
    for column in frame.columns:
        values = frame[column].dropna().astype(str).str.strip()
        unique_values = sorted(value for value in values.unique() if value != "")
        if len(unique_values) <= 120:
            distinct[column] = unique_values

    return {
        "name":name,
        "url":url,
        "bytes":len(response.content),
        "rows":int(len(frame)),
        "columns":list(frame.columns),
        "distinct_values":distinct,
        "sample_rows":frame.head(20).where(pd.notna(frame),None).to_dict(orient="records"),
    }


# ============================================================
# 04. DOWNLOAD AND WRITE INSPECTION OUTPUT
# ============================================================
def main() -> None:
    output = {
        "publication":"Hospital Admitted Patient Care Activity, 2024-25",
        "files":[
            inspect_workbook("summary",FILES["summary"]),
            inspect_workbook("icb",FILES["icb"]),
            inspect_workbook("provider",FILES["provider"]),
            inspect_csv("provider_level_analysis",FILES["provider_level_analysis"]),
        ],
    }

    OUTPUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output,indent=2,ensure_ascii=False),encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
