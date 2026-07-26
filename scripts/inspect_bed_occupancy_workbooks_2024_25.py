# ============================================================
# 00. INSPECT THE FOUR OFFICIAL 2024-25 OVERNIGHT-BED WORKBOOKS
# ============================================================
# Downloads each quarterly KH03 workbook and records its sheets and representative
# rows. Q1 and Q3 use the latest revised files currently linked by NHS England.

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any

import openpyxl
import requests


FILES = {
    "Q1": "https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2024/11/Beds-Open-Overnight-Web_File-Q1-2024-25-revised.xlsx",
    "Q2": "https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2024/11/Beds-Open-Overnight-Web_File-Q2-2024-25.xlsx",
    "Q3": "https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2025/11/Beds-Open-Overnight-Web_File-Q3-2024-25-revised.xlsx",
    "Q4": "https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2025/05/Beds-Open-Overnight-Web_File-Q4-2024-25.xlsx",
}
OUTPUT_PATH = Path("public-data/bed-occupancy-workbooks-2024-25-inspection.json")


def clean(value: Any) -> Any:
    if isinstance(value,(str,int,float,bool)) or value is None:
        return value
    return str(value)


def main() -> None:
    output = {"files":{}}

    for quarter,url in FILES.items():
        response = requests.get(url,timeout=300,headers={"User-Agent":"Mozilla/5.0"})
        response.raise_for_status()
        workbook = openpyxl.load_workbook(io.BytesIO(response.content),data_only=True,read_only=True)
        sheets = []

        for sheet in workbook.worksheets:
            preview = []
            row_numbers = list(range(1,min(sheet.max_row,35) + 1))
            if sheet.max_row > 35:
                row_numbers += list(range(max(36,sheet.max_row - 12),sheet.max_row + 1))
            for row_number in row_numbers:
                values = [clean(cell.value) for cell in sheet[row_number]]
                if any(value not in (None,"") for value in values):
                    preview.append({"row":row_number,"values":values})

            sheets.append({
                "name":sheet.title,
                "max_row":sheet.max_row,
                "max_column":sheet.max_column,
                "preview":preview,
            })

        output["files"][quarter] = {
            "url":url,
            "bytes":len(response.content),
            "sheet_names":workbook.sheetnames,
            "sheets":sheets,
        }

    OUTPUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output,indent=2,ensure_ascii=False),encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
