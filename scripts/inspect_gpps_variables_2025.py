# ============================================================
# 00. INSPECT GPPS 2025 REPORTING VARIABLES AND CANDIDATE METRICS
# ============================================================

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any
from urllib.parse import quote

import openpyxl
import pandas as pd
import requests


BASE = "https://www.gp-patient.co.uk/FileDownload/Download?fileRedirect="
FILES = {
    "national":"2025/survey-results/national-results/national-data-csv/GPPS_2025_National_data_(weighted)_(csv)_PUBLIC.csv",
    "ics":"2025/survey-results/ics-results/ics-data-csv/GPPS_2025_ICS_data_(weighted)_(csv)_PUBLIC.csv",
    "practice":"2025/survey-results/practice-results/practice-data-csv/GPPS_2025_Practice_data_(weighted)_(csv)_PUBLIC.csv",
}
VARIABLES_PATH = "2025/survey-results/supporting-documents/list-of-reporting-variables-csv-files/GPPS_2025_List_of_reporting_variables_PUBLIC.xlsx"
OUTPUT_PATH = Path("public-data/gpps-2025-variable-inspection.json")


def source_url(path: str) -> str:
    return BASE + quote(path,safe="")


def download(path: str) -> bytes:
    response = requests.get(source_url(path),timeout=300)
    response.raise_for_status()
    return response.content


def scalar(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value,"item"):
        return value.item()
    return value


def csv_profile(label: str,path: str) -> dict[str,Any]:
    frame = pd.read_csv(io.BytesIO(download(path)),dtype=str,low_memory=False,encoding="utf-8-sig")
    national = frame.iloc[0].to_dict() if len(frame) else {}
    stems = (
        "localgpservicesphone","localgpserviceswebsite","localgpservicesapp",
        "localgpservicesreception","localgpservicesprefhp","localgpservicesprefhpsee",
        "gpcontactoverall","gpcontacthandlerequest","gpcontactunsuccessful",
        "lastgpapptlisten","lastgpapptcare","lastgpapptconfidence",
        "overallgp","supportmanage","outofhours",
    )
    candidate_columns = [column for column in frame.columns if any(stem in column.lower() for stem in stems)]
    return {
        "label":label,
        "rows":len(frame),
        "column_count":len(frame.columns),
        "identifier_columns":list(frame.columns[:12]),
        "first_identifier_values":{column:scalar(national.get(column)) for column in frame.columns[:12]},
        "candidate_columns":[
            {"column":column,"first_value":scalar(national.get(column))}
            for column in candidate_columns
        ],
    }


def workbook_profile() -> dict[str,Any]:
    content = download(VARIABLES_PATH)
    workbook = openpyxl.load_workbook(io.BytesIO(content),read_only=True,data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        rows = []
        for row_number,row in enumerate(sheet.iter_rows(values_only=True),start=1):
            values = [scalar(value) for value in row]
            if any(value not in (None,"") for value in values):
                rows.append({"row":row_number,"values":values[:12]})
            if len(rows) >= 80:
                break
        sheets.append({
            "name":sheet.title,
            "max_row":sheet.max_row,
            "max_column":sheet.max_column,
            "preview":rows,
        })
    return {"bytes":len(content),"sheets":sheets}


def main() -> None:
    output = {
        "csv_files":[csv_profile(label,path) for label,path in FILES.items()],
        "reporting_variables":workbook_profile(),
    }
    OUTPUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output,indent=2,ensure_ascii=False),encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
