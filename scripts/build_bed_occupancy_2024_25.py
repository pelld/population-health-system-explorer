# ============================================================
# 00. PUBLIC BED AVAILABILITY AND OCCUPANCY — 2024-25
# ============================================================
# Downloads the four official KH03 overnight-bed workbooks. Quarterly values are
# direct published average daily beds. The annual view is a day-weighted average
# of those four published quarters; monthly or quarterly stocks are never summed.

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any

import openpyxl
import requests


PUBLICATION_URL = "https://www.england.nhs.uk/statistics/statistical-work-areas/bed-availability-and-occupancy/bed-data-overnight/"
FILES = {
    "Q1": {"url":"https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2024/11/Beds-Open-Overnight-Web_File-Q1-2024-25-revised.xlsx","days":91,"period":"April to June 2024"},
    "Q2": {"url":"https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2024/11/Beds-Open-Overnight-Web_File-Q2-2024-25.xlsx","days":92,"period":"July to September 2024"},
    "Q3": {"url":"https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2025/11/Beds-Open-Overnight-Web_File-Q3-2024-25-revised.xlsx","days":92,"period":"October to December 2024"},
    "Q4": {"url":"https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2025/05/Beds-Open-Overnight-Web_File-Q4-2024-25.xlsx","days":90,"period":"January to March 2025"},
}
OUTPUT_PATH = Path("public-data/bed-occupancy-2024-25.json")
PERIOD = "2024-25"
SECTORS = ["total","general_acute","learning_disabilities","maternity","mental_illness"]


# ============================================================
# 01. GENERIC CLEANING AND CALCULATION HELPERS
# ============================================================
def safe_float(value: Any,digits: int | None = None) -> float | None:
    if value in (None,"","-","*",":"):
        return None
    try:
        number = float(value)
    except (TypeError,ValueError):
        return None
    return round(number,digits) if digits is not None else number


def percentage(numerator: float | None,denominator: float | None) -> float | None:
    if numerator is None or denominator in (None,0):
        return None
    return round((numerator / denominator) * 100,2)


def clean_name(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    words = text.title().split()
    replacements = {"Nhs":"NHS","Cic":"CIC","Ltd":"Ltd","Hca":"HCA","Uk":"UK"}
    lower_words = {"And","Of","The","On","In","For","At"}
    output = []
    for index,word in enumerate(words):
        word = replacements.get(word,word)
        if index and word in lower_words:
            word = word.lower()
        output.append(word)
    return " ".join(output)


def row_values(sheet,row_number: int) -> list[Any]:
    return [cell.value for cell in sheet[row_number]]


def sector_values(row: list[Any]) -> dict[str,dict[str,float | None]]:
    available_indices = [6,7,8,9,10]
    occupied_indices = [12,13,14,15,16]
    output = {}
    for sector,available_index,occupied_index in zip(SECTORS,available_indices,occupied_indices):
        available = safe_float(row[available_index])
        occupied = safe_float(row[occupied_index])
        output[sector] = {
            "available":available,
            "occupied":occupied,
            "occupancy_percent":percentage(occupied,available),
        }
    return output


def weighted_average(quarter_values: dict[str,float | None]) -> float | None:
    weighted_total = 0.0
    total_days = 0
    for quarter,value in quarter_values.items():
        if value is None:
            continue
        days = int(FILES[quarter]["days"])
        weighted_total += value * days
        total_days += days
    return None if not total_days else round(weighted_total / total_days,2)


# ============================================================
# 02. READ DIRECT PUBLISHED QUARTERLY ROWS
# ============================================================
def download_workbook(url: str):
    response = requests.get(url,timeout=300,headers={"User-Agent":"Mozilla/5.0"})
    response.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(response.content),data_only=True,read_only=True)


def parse_quarter(quarter: str,metadata: dict[str,Any]) -> tuple[dict[str,Any],list[dict[str,Any]]]:
    workbook = download_workbook(str(metadata["url"]))
    sheet = workbook["NHS Trust by Sector"]

    england_row = row_values(sheet,16)
    england = {
        "quarter":quarter,
        "period":metadata["period"],
        "days":metadata["days"],
        "sectors":sector_values(england_row),
    }

    providers = []
    for row_number in range(18,sheet.max_row + 1):
        row = row_values(sheet,row_number)
        code = "" if len(row) <= 4 or row[4] in (None,"") else str(row[4]).strip()
        name = "" if len(row) <= 5 or row[5] in (None,"") else str(row[5]).strip()
        if not code or not name:
            continue
        providers.append({
            "code":code,
            "name":clean_name(name),
            "region_code":None if row[3] in (None,"") else str(row[3]).strip(),
            "quarter":quarter,
            "period":metadata["period"],
            "days":metadata["days"],
            "sectors":sector_values(row),
        })

    if len(providers) < 150:
        raise RuntimeError(f"{quarter}: expected at least 150 provider rows; found {len(providers)}")

    # Each published occupancy percentage must reconcile to occupied / available.
    for sector in SECTORS:
        published = safe_float(england_row[{"total":18,"general_acute":19,"learning_disabilities":20,"maternity":21,"mental_illness":22}[sector]])
        derived = england["sectors"][sector]["occupancy_percent"]
        if published is not None and derived is not None and abs((published * 100) - derived) > 0.02:
            raise RuntimeError(f"{quarter} {sector}: published occupancy did not reconcile ({published}; {derived}).")

    return england,providers


# ============================================================
# 03. BUILD ANNUAL DAY-WEIGHTED ENGLAND AND PROVIDER VIEWS
# ============================================================
def annual_sectors(quarter_records: dict[str,dict[str,Any]]) -> dict[str,dict[str,float | None]]:
    output = {}
    for sector in SECTORS:
        available = weighted_average({quarter:record["sectors"][sector]["available"] for quarter,record in quarter_records.items()})
        occupied = weighted_average({quarter:record["sectors"][sector]["occupied"] for quarter,record in quarter_records.items()})
        output[sector] = {
            "available":available,
            "occupied":occupied,
            "occupancy_percent":percentage(occupied,available),
        }
    return output


def metric(label: str,display: str,**values: Any) -> dict[str,Any]:
    return {"label":label,"display":display,**values}


def build_geography(code: str,name: str,geography_type: str,quarter_records: dict[str,dict[str,Any]]) -> dict[str,Any]:
    annual = annual_sectors(quarter_records)
    general_acute = annual["general_acute"]
    return {
        "code":code,
        "name":name,
        "type":geography_type,
        "period":PERIOD,
        "metrics":{
            "available-overnight-beds":metric(
                "Average daily general and acute beds available overnight",
                "beds",
                beds=general_acute["available"],
            ),
            "occupied-overnight-beds":metric(
                "Average daily general and acute beds occupied overnight",
                "beds",
                beds=general_acute["occupied"],
            ),
            "overnight-bed-occupancy":metric(
                "General and acute overnight-bed occupancy",
                "percent",
                percent=general_acute["occupancy_percent"],
                occupied=general_acute["occupied"],
                available=general_acute["available"],
            ),
        },
        "sectors":annual,
        "quarters":quarter_records,
        "quarters_reported":len(quarter_records),
    }


# ============================================================
# 04. VALIDATE AND WRITE THE COMPACT PUBLIC FILE
# ============================================================
def main() -> None:
    england_quarters = {}
    provider_quarters: dict[str,dict[str,Any]] = {}
    provider_names: dict[str,str] = {}

    for quarter,metadata in FILES.items():
        england,providers = parse_quarter(quarter,metadata)
        england_quarters[quarter] = england
        for provider in providers:
            code = provider["code"]
            provider_quarters.setdefault(code,{})[quarter] = provider
            provider_names[code] = provider["name"]

    england = build_geography("ENGLAND","England","National",england_quarters)
    providers = [
        build_geography(code,provider_names[code],"Provider",quarters)
        for code,quarters in provider_quarters.items()
        if any((record["sectors"]["total"]["available"] or 0) > 0 for record in quarters.values())
    ]
    providers.sort(key=lambda item:item["name"])

    if set(england_quarters) != set(FILES):
        raise RuntimeError("Not all four published quarters were included.")
    if len(providers) < 150:
        raise RuntimeError(f"Expected at least 150 providers with overnight beds; found {len(providers)}")

    # Hard validation of the direct published Q1 England row. The remaining exact
    # quarter values are also retained in the output and are checked internally.
    q1_ga = england_quarters["Q1"]["sectors"]["general_acute"]
    if round(q1_ga["available"] or 0,6) != round(105644.79120879123,6):
        raise RuntimeError(f"Q1 published England general and acute available beds changed: {q1_ga['available']}")
    if round(q1_ga["occupied"] or 0,6) != round(96361.15384615384,6):
        raise RuntimeError(f"Q1 published England general and acute occupied beds changed: {q1_ga['occupied']}")

    output = {
        "publication":"Bed Availability and Occupancy — Overnight Beds, 2024-25",
        "period":PERIOD,
        "source":"NHS England KH03 quarterly collection",
        "source_url":PUBLICATION_URL,
        "source_files":{quarter:metadata["url"] for quarter,metadata in FILES.items()},
        "provider_count":len(providers),
        "england":england,
        "providers":providers,
        "method":{
            "quarterly":"Direct published average daily available and occupied overnight-bed rows from each quarterly workbook.",
            "annual":"Day-weighted average of Q1 to Q4 using 91, 92, 92 and 90 calendar days. Occupancy is annual occupied bed-days divided by annual available bed-days.",
            "geography":"Provider geography only; no ICB attribution is inferred.",
        },
        "notes":[
            "General and acute beds exclude maternity and mental-health sectors; the full sector profile remains available in the details panel.",
            "Available means beds open for use during the reporting period, expressed as an average daily number.",
            "Occupancy is not case-mix adjusted and a high rate does not by itself establish unsafe care.",
            "Some quarterly returns were estimated or revised; the source workbooks' data-quality notes remain authoritative.",
        ],
    }

    OUTPUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output,indent=2,ensure_ascii=False),encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}: {len(providers)} providers; England occupancy {england['metrics']['overnight-bed-occupancy']['percent']}%")
    for quarter in FILES:
        ga = england_quarters[quarter]["sectors"]["general_acute"]
        print(quarter,ga["available"],ga["occupied"],ga["occupancy_percent"])


if __name__ == "__main__":
    main()
