# ============================================================
# 00. FINAL 2024-25 HES ADMITTED PATIENT CARE
# ============================================================
# Downloads the official final annual provider, ICB-responsibility and provider-
# level-analysis files. England totals are direct published totals. The ICB file
# publishes sub-ICB commissioner rows named by parent ICB; these are grouped to the
# 42 current ICBs without imputing locally suppressed values.

from __future__ import annotations

import io
import json
import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import requests


PUBLICATION_URL = "https://digital.nhs.uk/data-and-information/publications/statistical/hospital-admitted-patient-care-activity/2024-25"
ICB_URL = "https://files.digital.nhs.uk/B0/CC3BB5/hosp-epis-stat-admi-icb-resp-2024-25-tab.xlsx"
PROVIDER_URL = "https://files.digital.nhs.uk/DF/EE7338/hosp-epis-stat-admi-hosp-prov-2024-25-tab.xlsx"
PROVIDER_ANALYSIS_URL = "https://files.digital.nhs.uk/3C/145B13/hosp-epis-stat-admi-pla-2024-25.csv"
ICB_REFERENCE_PATH = Path("public-data/community-bed-audit-2026.json")
OUTPUT_PATH = Path("public-data/hes-apc-2024-25.json")
PERIOD = "2024-25"


# ============================================================
# 01. GENERIC CLEANING HELPERS
# ============================================================
def safe_int(value: Any) -> int | None:
    if value in (None,"","*",":","-"):
        return None
    try:
        return int(round(float(value)))
    except (TypeError,ValueError):
        return None


def safe_float(value: Any,digits: int = 2) -> float | None:
    if value in (None,"","*",":","-"):
        return None
    try:
        return round(float(value),digits)
    except (TypeError,ValueError):
        return None


def percentage(numerator: int | None,denominator: int | None) -> float | None:
    if numerator is None or denominator in (None,0):
        return None
    return safe_float((numerator / denominator) * 100)


def clean_name(value: Any,geography_type: str) -> str:
    text = "" if value is None else str(value).strip()
    if geography_type == "ICB":
        text = text.removeprefix("NHS ").removesuffix(" INTEGRATED CARE BOARD")
        text = re.sub(r" ICB - [A-Z0-9]+$","",text)
    words = text.title().split()
    replacements = {
        "Nhs":"NHS",
        "Icb":"ICB",
        "Hca":"HCA",
        "Cic":"CIC",
        "Ltd":"Ltd",
    }
    lower_words = {"And","Of","The","On","In","For","At"}
    output = []
    for index,word in enumerate(words):
        word = replacements.get(word,word)
        if index and word in lower_words:
            word = word.lower()
        output.append(word)
    return " ".join(output)


def normalise_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]","",value.lower())


def row_values(sheet,row_number: int) -> list[Any]:
    return [cell.value for cell in sheet[row_number]]


def metric(label: str,display: str,**values: Any) -> dict[str,Any]:
    return {"label":label,"display":display,**values}


# ============================================================
# 02. BUILD DIRECT PUBLISHED NATIONAL OR PROVIDER ROWS
# ============================================================
def build_geography(
    row: list[Any],
    geography_type: str,
    bed_days: int | None,
    code_override: str | None = None,
    name_override: str | None = None,
) -> dict[str,Any]:
    code = code_override or str(row[0]).strip()
    name = name_override or clean_name(row[1],geography_type)

    fces = safe_int(row[7])
    admissions = safe_int(row[8])
    emergency = safe_int(row[12])
    waiting_list = safe_int(row[13])
    planned = safe_int(row[14])
    other = safe_int(row[15])
    mean_wait = safe_float(row[16])
    median_wait = safe_float(row[17])
    mean_los = safe_float(row[18])
    median_los = safe_float(row[19])
    mean_age = safe_float(row[20])

    return geography_record(
        code=code,
        name=name,
        geography_type=geography_type,
        fces=fces,
        admissions=admissions,
        emergency=emergency,
        waiting_list=waiting_list,
        planned=planned,
        other=other,
        bed_days=bed_days,
        mean_wait=mean_wait,
        median_wait=median_wait,
        mean_los=mean_los,
        median_los=median_los,
        mean_age=mean_age,
    )


def geography_record(
    code: str,
    name: str,
    geography_type: str,
    fces: int | None,
    admissions: int | None,
    emergency: int | None,
    waiting_list: int | None,
    planned: int | None,
    other: int | None,
    bed_days: int | None,
    mean_wait: float | None,
    median_wait: float | None,
    mean_los: float | None,
    median_los: float | None,
    mean_age: float | None,
    aggregation: dict[str,Any] | None = None,
) -> dict[str,Any]:
    output = {
        "code":code,
        "name":name,
        "type":geography_type,
        "period":PERIOD,
        "metrics":{
            "emergency-admission":metric(
                "Emergency finished admission episodes",
                "count",
                count=emergency,
                percent=percentage(emergency,admissions),
                denominator=admissions,
                denominator_label="all finished admission episodes",
            ),
            "hes-bed-days":metric(
                "All admitted-patient finished consultant episode bed-days",
                "count",
                count=bed_days,
                percent=None,
            ),
            "hes-mean-los":metric(
                "Mean admitted spell length of stay",
                "days",
                days=mean_los,
                count=None,
                percent=None,
            ),
        },
        "activity":{
            "finished_consultant_episodes":fces,
            "finished_admission_episodes":admissions,
            "emergency_admissions":emergency,
            "waiting_list_admissions":waiting_list,
            "planned_admissions":planned,
            "other_admission_method":other,
            "emergency_share_percent":percentage(emergency,admissions),
            "bed_days":bed_days,
        },
        "duration":{
            "mean_time_waited_days":mean_wait,
            "median_time_waited_days":median_wait,
            "mean_length_of_stay_days":mean_los,
            "median_length_of_stay_days":median_los,
        },
        "mean_age":mean_age,
    }
    if aggregation:
        output["aggregation"] = aggregation
    return output


# ============================================================
# 03. AGGREGATE PUBLISHED SUB-ICB COMMISSIONER ROWS
# ============================================================
def load_icb_reference() -> dict[str,dict[str,str]]:
    if not ICB_REFERENCE_PATH.exists():
        raise RuntimeError(f"ICB reference file was not found: {ICB_REFERENCE_PATH}")
    data = json.loads(ICB_REFERENCE_PATH.read_text(encoding="utf-8"))
    return {
        normalise_name(item["name"]):{"code":item["code"],"name":item["name"]}
        for item in data.get("icbs",[])
    }


def aggregate_icbs(sheet) -> list[dict[str,Any]]:
    reference = load_icb_reference()
    fields = {
        "fces":7,
        "admissions":8,
        "emergency":12,
        "waiting_list":13,
        "planned":14,
        "other":15,
    }
    groups: dict[str,dict[str,Any]] = {}
    unmatched = set()

    for row_number in range(18,sheet.max_row + 1):
        row = row_values(sheet,row_number)
        raw_name = "" if row[1] in (None,"") else str(row[1]).strip()
        match = re.match(r"^NHS (.+?) ICB - [A-Z0-9]+$",raw_name,re.IGNORECASE)
        if not match:
            continue

        parent_name = clean_name(match.group(1),"ICB")
        reference_item = reference.get(normalise_name(parent_name))
        if not reference_item:
            unmatched.add(parent_name)
            continue

        code = reference_item["code"]
        group = groups.setdefault(code,{
            "code":code,
            "name":reference_item["name"],
            "values":{field:0 for field in fields},
            "suppressed":{field:0 for field in fields},
            "sub_icb_rows":0,
        })
        group["sub_icb_rows"] += 1

        for field,column_index in fields.items():
            value = safe_int(row[column_index])
            if value is None:
                if row[column_index] in ("*",":"):
                    group["suppressed"][field] += 1
            else:
                group["values"][field] += value

    if unmatched:
        raise RuntimeError(f"Could not map HES parent ICB names: {sorted(unmatched)}")

    output = []
    for group in groups.values():
        values = group["values"]
        output.append(geography_record(
            code=group["code"],
            name=group["name"],
            geography_type="ICB",
            fces=values["fces"],
            admissions=values["admissions"],
            emergency=values["emergency"],
            waiting_list=values["waiting_list"],
            planned=values["planned"],
            other=values["other"],
            bed_days=None,
            mean_wait=None,
            median_wait=None,
            mean_los=None,
            median_los=None,
            mean_age=None,
            aggregation={
                "method":"Sum of the published sub-ICB commissioner rows named for this parent ICB.",
                "sub_icb_rows":group["sub_icb_rows"],
                "suppressed_components":group["suppressed"],
                "warning":"Suppressed component values are not imputed; displayed ICB sums may therefore be slightly below the underlying total.",
            },
        ))

    output.sort(key=lambda item:item["name"])
    return output


# ============================================================
# 04. DOWNLOAD SOURCE FILES AND BED-DAY LOOKUPS
# ============================================================
def download_workbook(url: str):
    response = requests.get(url,timeout=300)
    response.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(response.content),read_only=True,data_only=True)


def bed_day_lookups() -> tuple[int,dict[str,int | None]]:
    response = requests.get(PROVIDER_ANALYSIS_URL,timeout=300)
    response.raise_for_status()
    frame = pd.read_csv(io.BytesIO(response.content),dtype=str,low_memory=False,encoding="utf-8-sig")
    rows = frame.loc[frame["MEASURE"].eq("FCE_BED_DAYS")].copy()

    england_rows = rows.loc[rows["ORG_LEVEL"].eq("ENGLAND")]
    if len(england_rows) != 1:
        raise RuntimeError(f"Expected one England FCE_BED_DAYS row; found {len(england_rows)}")
    england = safe_int(england_rows.iloc[0]["MEASURE_VALUE"])
    if england is None:
        raise RuntimeError("England FCE_BED_DAYS value was missing.")

    providers = {
        str(row["ORG_CODE"]).strip():safe_int(row["MEASURE_VALUE"])
        for _,row in rows.loc[rows["ORG_LEVEL"].eq("PROVIDER")].iterrows()
        if str(row["ORG_CODE"]).strip()
    }
    return england,providers


# ============================================================
# 05. BUILD AND VALIDATE THE COMPACT PUBLIC FILE
# ============================================================
def main() -> None:
    icb_workbook = download_workbook(ICB_URL)
    provider_workbook = download_workbook(PROVIDER_URL)
    icb_sheet = icb_workbook["ICB of Responsibility - Supp"]
    provider_sheet = provider_workbook["Hospital Providers"]
    england_bed_days,provider_bed_days = bed_day_lookups()

    provider_total_row = row_values(provider_sheet,17)
    icb_total_row = row_values(icb_sheet,17)
    england = build_geography(
        provider_total_row,
        geography_type="National",
        bed_days=england_bed_days,
        code_override="ENGLAND",
        name_override="England",
    )
    icbs = aggregate_icbs(icb_sheet)

    all_provider_rows = []
    for row_number in range(18,provider_sheet.max_row + 1):
        row = row_values(provider_sheet,row_number)
        code = "" if row[0] in (None,"") else str(row[0]).strip()
        name = "" if row[1] in (None,"") else str(row[1]).strip()
        admissions = safe_int(row[8])
        if not code or not name or admissions is None or "COMMISSIONING REGION" in name.upper():
            continue
        all_provider_rows.append(build_geography(row,"Provider",provider_bed_days.get(code)))

    providers = [item for item in all_provider_rows if (item["activity"]["emergency_admissions"] or 0) > 0]
    providers.sort(key=lambda item:item["name"])

    expected = {
        "finished_consultant_episodes":22555615,
        "finished_admission_episodes":18468856,
        "emergency_admissions":6614976,
        "bed_days":48360197,
        "mean_length_of_stay_days":4.69,
    }
    actual = {
        "finished_consultant_episodes":england["activity"]["finished_consultant_episodes"],
        "finished_admission_episodes":england["activity"]["finished_admission_episodes"],
        "emergency_admissions":england["activity"]["emergency_admissions"],
        "bed_days":england["activity"]["bed_days"],
        "mean_length_of_stay_days":england["duration"]["mean_length_of_stay_days"],
    }

    icb_total_check = build_geography(
        icb_total_row,
        geography_type="National",
        bed_days=england_bed_days,
        code_override="ENGLAND",
        name_override="England",
    )

    if actual != expected:
        raise RuntimeError(f"Published England HES totals did not reconcile: expected {expected}; found {actual}")
    if icb_total_check["activity"]["emergency_admissions"] != expected["emergency_admissions"]:
        raise RuntimeError("The ICB workbook total did not match the published emergency-admission total.")
    if len(icbs) != 42:
        raise RuntimeError(f"Expected 42 parent ICB aggregations; found {len(icbs)}")
    if len(providers) < 100:
        raise RuntimeError(f"Expected more than 100 providers with emergency admissions; found {len(providers)}")

    output = {
        "publication":"Hospital Admitted Patient Care Activity, 2024-25",
        "period":PERIOD,
        "source":"Hospital Episode Statistics — Admitted Patient Care",
        "source_url":PUBLICATION_URL,
        "source_files":{
            "icb":ICB_URL,
            "provider":PROVIDER_URL,
            "provider_level_analysis":PROVIDER_ANALYSIS_URL,
        },
        "icb_count":len(icbs),
        "provider_count":len(providers),
        "all_provider_row_count":len(all_provider_rows),
        "england":england,
        "icbs":icbs,
        "providers":providers,
        "coverage":{
            "recorded_icb_emergency_admissions":sum(item["activity"]["emergency_admissions"] or 0 for item in icbs),
            "recorded_provider_emergency_admissions":sum(item["activity"]["emergency_admissions"] or 0 for item in all_provider_rows),
            "recorded_provider_bed_days":sum(item["activity"]["bed_days"] or 0 for item in all_provider_rows),
        },
        "method":{
            "england":"Direct published Total rows; validated against the final annual summary and provider-level-analysis files.",
            "icb":"Sum of published sub-ICB commissioner rows grouped by the parent ICB named in the workbook; current Q codes are taken from the repository's official 42-ICB reference.",
            "provider":"Direct hospital-provider rows. The selector retains providers with recorded emergency admissions.",
            "bed_days":"FCE_BED_DAYS is published for England and providers. It covers all admitted patient care, not emergency activity alone, and is not available at ICB level in this public provider-analysis file.",
        },
        "notes":[
            "England totals are exact published values. Lower-geography counts can be rounded or suppressed, so local rows are not imputed to force reconciliation.",
            "ICB means ICB of responsibility: a commissioning geography, not necessarily residence or treatment location.",
            "Finished admission episodes count admissions; finished consultant episodes count periods under one consultant and are not unique patients.",
            "Mean length of stay is the published admitted-spell mean and is not case-mix adjusted.",
            "The publication notes that approximately 860 Bolton NHS Foundation Trust records may relate to virtual ward activity included in error.",
        ],
    }

    OUTPUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output,indent=2,ensure_ascii=False),encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}: {len(icbs)} ICBs and {len(providers)} emergency providers")


if __name__ == "__main__":
    main()
