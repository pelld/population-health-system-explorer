# ============================================================
# 00. INSPECT THE PUBLIC 2024-25 UCR WORKBOOK
# ============================================================
# Downloads the complete NHS England 2-hour Urgent Community Response workbook
# and records enough workbook structure to build a reproducible ICB/provider
# extract without guessing sheet names, header rows or month columns.

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


SOURCE_URL = "https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2025/06/2-hour-Urgent-Community-Response-Performance-Metrics_2024-25.xlsx"
OUTPUT_PATH = Path("public-data/ucr-2024-25-inspection.json")
MAX_ROWS = 35
MAX_COLUMNS = 24


def serialise(value: Any) -> Any:
    if value is None or isinstance(value,(str,int,float,bool)):
        return value
    return str(value)


def main() -> None:
    response = requests.get(SOURCE_URL,timeout=300)
    response.raise_for_status()

    workbook = load_workbook(io.BytesIO(response.content),read_only=True,data_only=True)
    sheets: list[dict[str,Any]] = []

    for worksheet in workbook.worksheets:
        preview = []
        for row_number,row in enumerate(worksheet.iter_rows(values_only=True),start=1):
            if row_number > MAX_ROWS:
                break
            values = [serialise(value) for value in row[:MAX_COLUMNS]]
            if any(value not in (None,"") for value in values):
                preview.append({"row":row_number,"values":values})

        sheets.append({
            "name":worksheet.title,
            "max_row":worksheet.max_row,
            "max_column":worksheet.max_column,
            "preview":preview,
        })

    output = {
        "source_url":SOURCE_URL,
        "bytes":len(response.content),
        "sheet_count":len(sheets),
        "sheets":sheets,
    }

    OUTPUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output,indent=2,ensure_ascii=False),encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH} with {len(sheets)} sheets")


if __name__ == "__main__":
    main()
