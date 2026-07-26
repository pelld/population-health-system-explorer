# ============================================================
# 00. INSPECT 2024-25 ACUTE DISCHARGE SITREP FILES
# ============================================================
# Downloads representative organisation CSVs around the 27 May 2024 definition
# change and the independently published national time-series workbook. The output
# records the structures needed for the final validated extractor.

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import requests


FILES = {
    "April 2024":"https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2026/07/Daily-discharge-sitrep-monthly-data-webfile-1-CSV-apr24.csv",
    "May 2024":"https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2026/07/Daily-discharge-sitrep-monthly-data-webfile-2-CSV-may24.csv",
    "June 2024":"https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2026/07/Daily-discharge-sitrep-monthly-data-webfile-3-CSV-jun24.csv",
    "March 2025":"https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2026/07/Daily-discharge-sitrep-monthly-data-webfile-12-CSV-mar25.csv",
}
TIMESERIES_URL = "https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2026/07/Daily-discharge-sitrep-timeseries-data-webfile-April2021-June2026.xlsx"
OUTPUT_PATH = Path("public-data/acute-discharge-2024-25-inspection.json")


def scalar(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value,"item"):
        return value.item()
    return value


def inspect_file(label: str,url: str) -> dict[str,Any]:
    response = requests.get(url,timeout=300)
    response.raise_for_status()
    frame = pd.read_csv(io.BytesIO(response.content),dtype=str,low_memory=False,encoding="utf-8-sig")

    dimensions: dict[str,list[str]] = {}
    for column in frame.columns:
        normalised = column.lower().replace("_"," ")
        if any(token in normalised for token in ("level","metric","measure","organisation type","org type","breakdown","category","period","date")):
            values = sorted({str(value).strip() for value in frame[column].dropna() if str(value).strip()})
            if len(values) <= 250:
                dimensions[column] = values

    national = frame.loc[frame["Level"].eq("National")].copy()
    national_metric_shape = []
    for metric,rows in national.groupby("Metric",dropna=False):
        national_metric_shape.append({
            "metric":str(metric),
            "rows":len(rows),
            "periods":rows["Period"].nunique(dropna=True),
            "first_period":rows["Period"].dropna().min() if rows["Period"].notna().any() else None,
            "last_period":rows["Period"].dropna().max() if rows["Period"].notna().any() else None,
            "metric_type":sorted(rows["Metric Type"].dropna().unique().tolist()) if "Metric Type" in rows else [],
            "metric_group":sorted(rows["Metric Group"].dropna().unique().tolist()) if "Metric Group" in rows else [],
        })

    samples = []
    for _,row in frame.head(30).iterrows():
        samples.append({column:scalar(row[column]) for column in frame.columns})

    return {
        "label":label,
        "url":url,
        "bytes":len(response.content),
        "rows":len(frame),
        "columns":list(frame.columns),
        "dimensions":dimensions,
        "national_metric_shape":national_metric_shape,
        "sample_rows":samples,
    }


def inspect_timeseries() -> dict[str,Any]:
    response = requests.get(TIMESERIES_URL,timeout=300)
    response.raise_for_status()
    workbook = openpyxl.load_workbook(io.BytesIO(response.content),read_only=True,data_only=True)

    sheets = []
    for sheet in workbook.worksheets:
        preview = []
        for row_number,row in enumerate(sheet.iter_rows(values_only=True),start=1):
            values = list(row)
            if any(value not in (None,"") for value in values):
                preview.append({"row":row_number,"values":values[:30]})
            if len(preview) >= 35:
                break
        sheets.append({
            "name":sheet.title,
            "max_row":sheet.max_row,
            "max_column":sheet.max_column,
            "preview":preview,
        })

    return {
        "url":TIMESERIES_URL,
        "bytes":len(response.content),
        "sheet_names":workbook.sheetnames,
        "sheets":sheets,
    }


def main() -> None:
    output = {
        "files":[inspect_file(label,url) for label,url in FILES.items()],
        "national_timeseries":inspect_timeseries(),
    }
    OUTPUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output,indent=2,ensure_ascii=False,default=str),encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
