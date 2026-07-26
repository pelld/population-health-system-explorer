# ============================================================
# 00. INSPECT THE PUBLIC COMMUNITY WAITING-LIST WORKBOOK
# ============================================================
# The March 2025 publication contains the corrected 2024-25 time series. This
# temporary inspection output records sheet names, dimensions and representative
# rows so the production build can use the published structure rather than
# guessed sheet or column names.

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any

import openpyxl
import requests


SOURCE_URL = "https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2025/05/Community-health-services-waiting-lists-2024-25-March.xlsx"
OUTPUT_PATH = Path("public-data/chs-waits-workbook-inspection.json")


def serialise(value: Any) -> Any:
    if value is None or isinstance(value,(str,int,float,bool)):
        return value
    if hasattr(value,"isoformat"):
        return value.isoformat()
    return str(value)


def main() -> None:
    response = requests.get(SOURCE_URL,timeout=300)
    response.raise_for_status()

    workbook = openpyxl.load_workbook(io.BytesIO(response.content),read_only=True,data_only=True)
    sheets = []

    for worksheet in workbook.worksheets:
        preview = []

        for row_number,row in enumerate(worksheet.iter_rows(min_row=1,max_row=min(worksheet.max_row,80),values_only=True),start=1):
            values = [serialise(value) for value in row[:35]]
            if any(value not in (None,"") for value in values):
                preview.append({"row":row_number,"values":values})
            if len(preview) >= 35:
                break

        sheets.append({
            "name":worksheet.title,
            "max_row":worksheet.max_row,
            "max_column":worksheet.max_column,
            "preview":preview,
        })

    output = {
        "source_url":SOURCE_URL,
        "bytes":len(response.content),
        "sheet_count":len(workbook.sheetnames),
        "sheet_names":workbook.sheetnames,
        "sheets":sheets,
    }

    OUTPUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output,indent=2,ensure_ascii=False),encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH} with {len(sheets)} sheets")


if __name__ == "__main__":
    main()
