# ============================================================
# 00. INSPECT THE 4 MARCH 2026 COMMUNITY BED AUDIT WORKBOOK
# ============================================================
# Downloads the official NHS England workbook and records sheet names, dimensions
# and representative rows. This temporary output is used to build a validated
# production extractor from the published summary tables.

from __future__ import annotations

import io
import json
from pathlib import Path

import openpyxl
import requests


SOURCE_URL = "https://www.england.nhs.uk/statistics/wp-content/uploads/sites/2/2026/06/Community-Bed-Audit-4th-March-2026-v2.xlsx"
OUTPUT_PATH = Path("public-data/community-bed-audit-2026-inspection.json")


def serialise(value):
    if value is None or isinstance(value,(str,int,float,bool)):
        return value
    return str(value)


def inspect_sheet(sheet) -> dict:
    preview = []
    for row_number,row in enumerate(sheet.iter_rows(min_row=1,max_row=min(sheet.max_row,80),values_only=True),start=1):
        values = [serialise(value) for value in row[:min(sheet.max_column,60)]]
        if any(value not in (None,"") for value in values):
            preview.append({"row":row_number,"values":values})

    return {
        "name":sheet.title,
        "max_row":sheet.max_row,
        "max_column":sheet.max_column,
        "preview":preview,
    }


def main() -> None:
    response = requests.get(SOURCE_URL,timeout=300)
    response.raise_for_status()

    workbook = openpyxl.load_workbook(io.BytesIO(response.content),data_only=True,read_only=True)
    output = {
        "source_url":SOURCE_URL,
        "workbook_bytes":len(response.content),
        "sheet_names":workbook.sheetnames,
        "sheets":[inspect_sheet(workbook[name]) for name in workbook.sheetnames],
    }

    OUTPUT_PATH.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output,indent=2,ensure_ascii=False),encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}: {len(workbook.sheetnames)} sheets")


if __name__ == "__main__":
    main()
